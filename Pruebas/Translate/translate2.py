import pandas as pd
from googletrans import Translator
from openpyxl import load_workbook, Workbook

translator = Translator()

def translateToEnglish(text):
    translated = translator.translate(text, src="es", dest="en")
    return translated.text

def translateToSpanish(text):
    translated = translator.translate(text, src="en", dest="es")
    return translated.text

#try: 
i=0
filename = "c:/Users/brana/Desktop/Python/documentos_pruebas/book1.xlsx"
destinationfile = "c:/Users/brana/Desktop/Python/documentos_pruebas/book1_traduccion.xlsx"
# Load the source Excel file
source_wb = load_workbook(filename)
source_ws = source_wb.active
# Create a new Excel file
destination_wb = load_workbook(filename)
destination_ws = destination_wb.active
# Delete the rows of the destination file
destination_ws.delete_rows(1,source_ws.max_row)
for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, values_only=True):
    print("Fila: ",i)
    if row[0] == '':
        destination_ws.append({1:row[0],2:'No message'}) #Empty cell
    else:
        destination_ws.append({1:row[0],2:translateToSpanish(row[0])}) #Translate to English
    i+=1
# Save the destination file
destination_wb.save(destinationfile)
#Close the files
destination_wb.close()
source_wb.close()
print('Translation complete!')

'''
except Exception as e:
    #Close the files
    destination_wb.close()
    source_wb.close()
    #Print the error
    print('An error ocurred:', e)
    print('Translation failed!')
    
'''