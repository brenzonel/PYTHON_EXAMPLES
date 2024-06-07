from googletrans import Translator
from openpyxl import load_workbook


def translateToEnglish(text):
    translator = Translator()
    translated = translator.translate(text, src="es", dest="en")
    return translated.text

def translateToSpanish(text):
    translator = Translator()
    translated = translator.translate(text, src="en", dest="es")
    return translated.text

try:
    #File paths
    filename = "Carpeta/Archivo.xlsx"
    destinationfile = "Carpeta/Archivo_traduccion.xlsx"

    #Load the source Excel file
    soruce_wb = load_workbook(filename)
    source_ws = soruce_wb.active

    #Create a new Excel file
    destination_wb = load_workbook(filename)
    destination_ws = destination_wb.active

    #Iteration through the sheets
    for sheet in destination_wb.sheetnames:
        destination_ws = destination_wb[sheet]
        #Iteration through the rows and columns
        for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, min_col=1, max_col=source_ws.max_column):
            for cell in row:
                s = str(cell.value)
                if s == '' or s == None or s == 'None' or s == 'CHUBB':
                    print('Empty cell')
                else:
                    print(s)
                    try:
                        s1 = translateToEnglish(s)
                    except Exception as e:
                        print('Error:', e)
                        s1 = s
                    cell.value = s1
                    print(cell.value)

    #Save the destination file
    destination_wb.save(destinationfile)
    #Close the files
    destination_wb.close()
    soruce_wb.close()

    print('Translation complete!')

except Exception as e:
    #Close the files
    destination_wb.close()
    soruce_wb.close()
    #Print the error
    print('An error ocurred:', e)
    print('Translation failed!')
